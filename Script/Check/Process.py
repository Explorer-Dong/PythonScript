# @Time   : 2023-12-26 16:27
# @File   : Process.py
# @Author : Mr_Dwj

import os
import openpyxl
import xlrd


class Process:
	
	def __init__(self, files_path: str) -> None:
		self.files_path = files_path
	
	
	def xls2xlsx(self) -> None:
		
		for file_name in os.listdir(self.files_path):
			if not file_name.endswith('.xls'): continue
			
			# 获取xls文件路径
			xls_file_path = os.path.join(self.files_path, file_name)
			
			# 使用xlrd库打开xls文件并读取数据
			xls_file = xlrd.open_workbook(xls_file_path)
			sheets = xls_file.sheet_names()
			data = {}
			for sheet_name in sheets:
				sheet = xls_file.sheet_by_name(sheet_name)
				rows = []
				for row_idx in range(sheet.nrows):
					row = []
					for col_idx in range(sheet.ncols):
						cell_value = sheet.cell_value(row_idx, col_idx)
						row.append(cell_value)
					rows.append(row)
				data[sheet_name] = rows
			
			# 使用openpyxl库将数据写入新的.xlsx文件
			new_file_name = file_name.replace('.xls', '.xlsx')
			new_file_path = os.path.join(self.files_path, new_file_name)
			workbook = openpyxl.Workbook()
			workbook.remove(workbook.active)
			for sheet_name, rows in data.items():
				sheet = workbook.create_sheet(sheet_name)
				for row_idx, row in enumerate(rows):
					for col_idx, cell_value in enumerate(row):
						sheet.cell(row=row_idx + 1, column=col_idx + 1, value=cell_value)
			workbook.save(new_file_path)
			
			# 删除原来的.xls文件
			os.remove(xls_file_path)
			
			# 打印提示信息
			print(f'{file_name}\t --> \t{new_file_name}.')
	
	
	def read(self) -> None:
		pass
	
	
	def write(self) -> None:
		pass
