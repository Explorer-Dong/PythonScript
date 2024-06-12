import os
import xlrd
import openpyxl
from openpyxl.styles import PatternFill


# 定义一个函数，用于将.xls文件转换为.xlsx文件
def convert_xls_to_xlsx(folder_path):
    for file_name in os.listdir(folder_path):
        # 判断文件是否为.xls文件
        if file_name.endswith('.xls'):
            file_path = os.path.join(folder_path, file_name)
            # 使用xlrd库打开.xls文件，并读取数据
            workbook = xlrd.open_workbook(file_path)
            sheets = workbook.sheet_names()
            data = {}
            for sheet_name in sheets:
                sheet = workbook.sheet_by_name(sheet_name)
                rows = []
                for row_idx in range(sheet.nrows):
                    row = []
                    for col_idx in range(sheet.ncols):
                        cell_value = sheet.cell_value(row_idx, col_idx)
                        row.append(cell_value)
                    rows.append(row)
                data[sheet_name] = rows
            # 使用openpyxl库将数据写入新的.xlsx文件
            new_file_name = file_name.replace('.xls', '.xlsx')
            new_file_path = os.path.join(folder_path, new_file_name)
            workbook = openpyxl.Workbook()
            workbook.remove(workbook.active)
            for sheet_name, rows in data.items():
                sheet = workbook.create_sheet(sheet_name)
                for row_idx, row in enumerate(rows):
                    for col_idx, cell_value in enumerate(row):
                        sheet.cell(row=row_idx + 1, column=col_idx + 1, value=cell_value)
            workbook.save(new_file_path)


# 定义一个函数，用于删除指定路径下的所有后缀为.xls的文件
def delete_xls_files(folder_path):
    for filename in os.listdir(folder_path):
        if filename.endswith('.xls'):
            os.remove(os.path.join(folder_path, filename))
            print(f'{filename} has been deleted.')


# 定义一个函数，用于匹配两个Excel表格，如果申请正确，则写入新的Excel，否则不写入并且在申请表中的用黄色标注出来
def compare_and_remove_rows(source_file, compare_file, start_row, column):
    # 打开申请文件
    source_wb = openpyxl.load_workbook(source_file)
    source_ws = source_wb.active

    # 获取最大行数和最大列数
    max_row = source_ws.max_row
    max_column = source_ws.max_column

    # 打开对照文件
    compare_wb = openpyxl.load_workbook(compare_file)
    compare_ws = compare_wb.active

    # 获取对照文件的单元格值列表，作为正确参照物
    compare_values = []
    for cell in compare_ws["A"]:  # 这个比较文件的起始列号是 对照表 的 学号 一栏
        compare_values.append(cell.value)

    # 存储匹配成功的行数据
    row_data_list = []

    # 逐行处理
    fill_color = PatternFill("solid", fgColor="FFFF00")  # 黄色背景填充
    for row_num in range(start_row, max_row + 1):
        value = source_ws[column + str(row_num)].value

        if value in compare_values:
            # 存储 对照表 中的一行数据
            row_data = []
            for column_num in range(1, compare_ws.max_column + 1):
                cell = compare_ws.cell(row=compare_values.index(value) + 1, column=column_num)
                row_data.append(cell.value)
            row_data_list.append(row_data)  # 将匹配成功的 对照表中的 行添加到新的表格中
        else:
            for cell in source_ws[row_num]:  # 在源文件中对匹配失败的行加上黄色背景填充
                cell.fill = fill_color

    # 保存新的文件
    match_wb = openpyxl.Workbook()
    match_ws = match_wb.active

    # 添加表头
    header_row = []
    for column_num in range(1, compare_ws.max_column + 1):
        header_row.append(compare_ws.cell(row=1, column=column_num).value)
    match_ws.append(header_row)

    # 将所有匹配成功的行数据添加到新的工作簿中
    for row_data in row_data_list:
        match_ws.append(row_data)

    # 保存新的文件
    file_name, ext = os.path.splitext(source_file)
    new_filename = file_name + '_正确申请名单' + '.xlsx'
    match_wb.save(filename=new_filename)

    # 保存源文件
    source_wb.save(filename=source_file)

    # 输出运行结果
    print(f"{source_file}处理完成！")


if __name__ == '__main__':
    folder_path = r'E:\coding_Python\office库\Excel志愿审核\文件集合'

    # 调用convert_xls_to_xlsx函数将文件夹中的.xls文件转换为.xlsx文件
    convert_xls_to_xlsx(folder_path)

    # 调用delete_xls_files函数将文件夹中的.xls文件删除
    delete_xls_files(folder_path)

    # 调用compare_and_remove_rows函数进行匹配人员名单
    source_path = r"E:\coding_Python\office库\Excel志愿审核\文件集合\申请表.xlsx"
    compare_path = r"E:\coding_Python\office库\Excel志愿审核\文件集合\对照表.xlsx"
    start_row = 2  # 这是申请表的开始匹配的行数
    column = "C"  # 这是申请表的开始匹配的列数
    compare_and_remove_rows(source_path, compare_path, start_row, column)
